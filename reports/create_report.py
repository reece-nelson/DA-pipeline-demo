import psycopg2
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series

def create_report():

    # grab gold data
    conn = psycopg2.connect(
        dbname="default_database",
        user="username",
        password="password",
        port="5432",
        host="postgresdb"
    )   
    cur = conn.cursor()
    query = """
    SELECT
        *
    FROM gold.summary_eia_nhl
    """
    df = pd.read_sql(query, conn)
    conn.close()

    df_graph = df[df["team_name"] != "no_team_found"]
    df_graph = df_graph[["period","team_name","avg_revenue"]]
    df_graph['avg_revenue'] = pd.to_numeric(df_graph['avg_revenue'])
    pivot_df = df_graph.pivot(index="period", columns="team_name", values="avg_revenue").reset_index()
    cols_to_check = [col for col in pivot_df.columns if col != "period"]
    cols_to_drop = [col for col in cols_to_check if (pivot_df[col] < 900).all()]
    pivot_df = pivot_df.drop(columns=cols_to_drop)

    
    # create excel
    excel_file = "output.xlsx"
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        pivot_df.to_excel(writer, sheet_name="gold_eia_nhl_data", index=False)

    wb = load_workbook(excel_file)
    ws = wb["gold_eia_nhl_data"]
    chart = LineChart()
    chart.title = "Average Electricity Revenue of EIA by NHL Team"
    chart.x_axis.title = "Period"
    chart.y_axis.title = "Average Revenue"


    for idx, col in enumerate(pivot_df.columns[1:], start=2):
        if col != "period":
            data = Reference(ws, min_col=idx, min_row=2, max_row=ws.max_row)
            series = Series(data, title=col)
            chart.series.append(series)
        else:
            pass


    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)    
    chart.set_categories(cats_ref)

    ws_chart = wb.create_sheet(title="LineChartRevenue")
    ws_chart.add_chart(chart, "B2")

    output_path = os.path.join("/app/reports", excel_file)
    wb.save(output_path)